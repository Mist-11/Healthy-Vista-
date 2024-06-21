from django.shortcuts import render, redirect
from django.http import JsonResponse, FileResponse, HttpResponse
from users import models
from django.urls import reverse
import datetime
from openpyxl import load_workbook
import os
from tool import send_code_email


def login_page(request):
    if request.session.get('is_login', None):
        return redirect(reverse('index'))
    else:
        return render(request, 'login.html')


def login(request):
    if request.method == "POST":
        username = request.POST.get('username', None)
        password = request.POST.get('password', None)
        if username and password:  # 确保用户名和密码都不为空
            try:
                user = models.User.objects.get(name=username)
            except models.User.DoesNotExist:
                return JsonResponse({"message": "未寻找到该账户，请重新输入。"})
            if user.password == password:
                request.session['is_login'] = True
                request.session['username'] = user.name
                request.session['sex'] = user.sex
                if user.identity == '用户':
                    request.session['identity'] = '用户'
                    return JsonResponse({'redirect_url': reverse('index')})
                elif user.identity == '管理员' or "系统管理员":
                    request.session['identity'] = user.identity
                    return JsonResponse({'redirect_url': reverse('backstageindex')})
                else:
                    return JsonResponse({"message": "账号异常，请联系管理员。"})
            else:
                return JsonResponse({"message": "密码错误，请重新输入。"})
        else:
            return JsonResponse({"message": "请输入完整。"})
    return JsonResponse({"message": "请不要使用非法请求。"})


def register_page(request):
    if request.session.get('is_login', None):
        return redirect(reverse('index'))
    else:
        return render(request, 'register.html')


def forget_page(request):
    if request.session.get('is_login', None):
        return redirect(reverse('index'))
    else:
        return render(request, 'forgetpassword.html')


def register(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        password = request.POST.get('password')
        email = request.POST.get('email')
        again_password = request.POST.get('again-password')
        sex = request.POST.get('sex')

        if name and password and email:
            if models.User.objects.filter(name=name).exists():
                return JsonResponse({"message": "该用户名已被注册，请修改用户名。"})
            if models.User.objects.filter(email=email).exists():
                return JsonResponse({"message": "该邮箱已被使用，请更换邮箱。"})
            if password != again_password:
                return JsonResponse({"message": "两次输入密码不一致。"})
            print(datetime.datetime.now())
            # 创建新用户
            models.User.objects.create(
                name=name,
                password=password,
                email=email,
                sex=sex,
                identity='用户',
                timestamp=datetime.datetime.now()
            )
            # 重定向到登录页面或其他适当的页面
            return JsonResponse({'redirect_url': reverse('login')})
        else:
            return JsonResponse({"message": "请输入完整。"})
    return JsonResponse({"message": "请不要使用非法请求。"})


def logout(request):
    request.session.flush()
    return redirect("login")


def index(request):
    if not request.session.get('is_login', None):
        return redirect(reverse('login'))
    else:
        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        username = request.session.get('username')
        identity = '用户'
        if request.session.get("identity") == "管理员" or '系统管理员':
            identity = request.session.get("identity")
        message = {
            'current_date': current_date,
            'username': username,
            "identity": identity,
        }
        return render(request, 'index.html', message)


def usercenter_page(request):
    if not request.session.get('is_login', None):
        return redirect(reverse('login'))
    else:
        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        username = request.session.get('username')
        records = models.SimpleRecord.objects.filter(name=username).order_by('-timestamp')
        message = {
            'current_date': current_date,
            'username': username,
            'records': records,
        }
        return render(request, 'usercenter.html', message)


def find_detail_record(request, username, timestamp):
    if not request.session.get('is_login', None):
        return redirect(reverse('login'))
    else:
        pdf_path = f"./AllPDF/{timestamp}_{username}.pdf"
        print(pdf_path)
        if os.path.exists(pdf_path):
            return FileResponse(open(pdf_path, 'rb'), content_type='application/pdf')
        else:
            return HttpResponse("PDF not found", status=404)


def help_page(request):
    if not request.session.get('is_login', None):
        return redirect(reverse('login'))
    else:
        name = request.session.get('username')
        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        message = {
            "current_date": current_date,
            'username': name,
        }

        return render(request, 'helpcenter.html', message)


def timestamptotime(timestamp):
    # timestamp是一个以毫秒为单位的时间戳，将毫秒转换为秒
    timestamp = float(timestamp) / 1000.0
    # 使用datetime.fromtimestamp来转换时间戳为datetime对象
    timestamp = datetime.datetime.fromtimestamp(timestamp)
    return timestamp.strftime("%Y%m%d%H%M%S")  # 格式化


def single_create(request):
    name = request.POST.get('name', None)
    password = request.POST.get('password', None)
    email = request.POST.get('email', None)
    sex = request.POST.get('sex', None)
    models.User.objects.create(
        name=name,
        password=password,
        email=email,
        sex=sex,
        identity='管理员',
        timestamp=datetime.datetime.now()
    )
    return JsonResponse({"message": "创建成功！"})


def batch_create(request):
    if 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            username, password, email, sex = row
            models.User.objects.create(
                name=username,
                password=password,
                email=email,
                sex=sex,
                identity='管理员',
                timestamp=datetime.datetime.now()
            )
        return JsonResponse({"message": "批量创建成功！"})
    else:
        return JsonResponse({'message': '文件上传失败', 'status': 'error'}, status=400)


def sendcode(request):
    email = request.POST.get('email', None)
    print(email)
    result = models.EmailVerifyRecord.objects.filter(email=email).last()
    if result is None:
        result = send_code_email(email, send_type="retrieve")
        if result:
            return JsonResponse({'message': "验证码发送成功！"})
        else:
            return JsonResponse({'message': "验证码发送失败，请重试！"})
    if (datetime.datetime.now() - result.send_time.replace(tzinfo=None)) < datetime.timedelta(minutes=10):
        return JsonResponse({'message': "验证码已发送，请稍等！"})
    else:
        result = send_code_email(email, send_type="retrieve")
        if result:
            return JsonResponse({'message': "验证码发送成功！"})
        else:
            return JsonResponse({'message': "验证码发送失败，请重试！"})


def receivecode(request):
    code = request.POST.get('code', None)
    name = request.POST.get('name', None)
    password = request.POST.get('password', None)
    passwordagain = request.POST.get('passwordagain', None)
    result = models.EmailVerifyRecord.objects.filter(code=code).last()
    if code == result.code:
        if (datetime.datetime.now() - result.send_time.replace(tzinfo=None)) < datetime.timedelta(minutes=10):
            user = models.User.objects.get(name=name)
            if user:
                if password == passwordagain:
                    user.password = password
                    user.save()
                    return JsonResponse({'message': "密码修改成功！"})
                else:
                    return JsonResponse({'message': "两次输入密码不一致！"})
            else:
                return JsonResponse({'message': "该用户不存在。"})
        else:
            return JsonResponse({'message': "该验证码已失效！"})
    else:
        return JsonResponse({'message': "验证码错误！"})
